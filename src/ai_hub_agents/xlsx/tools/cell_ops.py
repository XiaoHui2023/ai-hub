"""单元格读写工具。"""

from __future__ import annotations

from typing import Annotated, Optional

from langchain_core.tools import ToolException, tool
from langgraph.prebuilt import InjectedState
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from ._helpers import get_output_file, with_file_lock


@tool
@with_file_lock
def read_cells(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    range_str: str,
    data_only: bool = False,
) -> str:
    """读取输出文件中指定范围的单元格值。

    Args:
        sheet_name: 工作表名称
        range_str: 单元格范围，如 "A1", "A1:D10"
        data_only: True 时返回缓存的计算值而非公式
    """
    try:
        wb = load_workbook(get_output_file(state), data_only=data_only)
        ws = wb[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        rows = []
        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col,
        ):
            rows.append("\t".join(str(c.value) for c in row))
        wb.close()
        return "\n".join(rows) if rows else "(空范围)"
    except KeyError:
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    except Exception as e:
        raise ToolException(f"read_cells 失败: {e}") from e


@tool
@with_file_lock
def write_cells(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    cells: list[dict],
) -> str:
    """批量写入单元格值或公式到输出文件。

    Args:
        sheet_name: 工作表名称
        cells: 写入列表，每项为 {"cell": "A1", "value": ...}。
               value 以 '=' 开头视为公式，如 {"cell": "B10", "value": "=SUM(B2:B9)"}
    """
    output = get_output_file(state)
    try:
        wb = load_workbook(output)
        ws = wb[sheet_name]
        for item in cells:
            ws[item["cell"]] = item["value"]
        wb.save(output)
        return f"已写入 {len(cells)} 个单元格到 '{sheet_name}'"
    except KeyError:
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    except Exception as e:
        raise ToolException(f"write_cells 失败: {e}") from e


@tool
@with_file_lock
def write_range(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    start_cell: str,
    data: list[list],
    headers: Optional[list[str]] = None,
) -> str:
    """从指定起始单元格写入二维数据到输出文件。

    Args:
        sheet_name: 工作表名称
        start_cell: 起始单元格，如 "A1"
        data: 二维数组 [[row1_col1, row1_col2, ...], ...]
        headers: 可选列标题，写在 data 之前
    """
    from openpyxl.utils import column_index_from_string

    output = get_output_file(state)
    try:
        wb = load_workbook(output)
        ws = wb[sheet_name]

        col_letter = "".join(c for c in start_cell if c.isalpha())
        row_num = int("".join(c for c in start_cell if c.isdigit()))
        start_col = column_index_from_string(col_letter)

        cur_row = row_num
        if headers:
            for ci, h in enumerate(headers):
                ws.cell(row=cur_row, column=start_col + ci, value=h)
            cur_row += 1

        for row_data in data:
            for ci, val in enumerate(row_data):
                ws.cell(row=cur_row, column=start_col + ci, value=val)
            cur_row += 1

        wb.save(output)
        total = len(data) + (1 if headers else 0)
        cols = max((len(r) for r in data), default=0)
        if headers:
            cols = max(cols, len(headers))
        return f"已写入 {total} 行 × {cols} 列到 '{sheet_name}' {start_cell} 起"
    except KeyError:
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    except Exception as e:
        raise ToolException(f"write_range 失败: {e}") from e


@tool
@with_file_lock
def clear_range(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    range_str: str,
    clear_format: bool = False,
) -> str:
    """清除输出文件中指定范围的单元格值（可选同时清除格式）。

    Args:
        sheet_name: 工作表名称
        range_str: 单元格范围，如 "A1:D10"
        clear_format: 是否同时清除格式
    """
    from openpyxl.styles import DEFAULT_FONT, Alignment, PatternFill

    output = get_output_file(state)
    try:
        wb = load_workbook(output)
        ws = wb[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        count = 0
        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col,
        ):
            for cell in row:
                cell.value = None
                if clear_format:
                    cell.font = DEFAULT_FONT
                    cell.fill = PatternFill()
                    cell.alignment = Alignment()
                    cell.number_format = "General"
                count += 1
        wb.save(output)
        return f"已清除 {count} 个单元格 ({range_str})"
    except KeyError:
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    except Exception as e:
        raise ToolException(f"clear_range 失败: {e}") from e
