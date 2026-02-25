"""工作表管理工具。"""

from __future__ import annotations

from typing import Annotated, Optional

from langchain_core.tools import ToolException, tool
from langgraph.prebuilt import InjectedState
from openpyxl import load_workbook

from ._helpers import get_output_file, with_file_lock


def _load(state: dict):
    return load_workbook(get_output_file(state))


def _save(wb, state: dict):
    wb.save(get_output_file(state))


@tool
@with_file_lock
def list_sheets(state: Annotated[dict, InjectedState]) -> str:
    """列出输出文件中所有工作表的名称、行数和列数。"""
    wb = _load(state)
    parts = []
    for name in wb.sheetnames:
        ws = wb[name]
        parts.append(f"  {name}: {ws.max_row} 行 × {ws.max_column} 列")
    wb.close()
    return f"共 {len(parts)} 个工作表:\n" + "\n".join(parts)


@tool
@with_file_lock
def add_sheet(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    index: Optional[int] = None,
) -> str:
    """在输出文件中添加一个新工作表。

    Args:
        sheet_name: 新工作表名称
        index: 插入位置（0 表示最前面），不指定则追加到末尾
    """
    wb = _load(state)
    if sheet_name in wb.sheetnames:
        wb.close()
        raise ToolException(f"工作表 '{sheet_name}' 已存在")
    wb.create_sheet(sheet_name, index)
    _save(wb, state)
    return f"已添加工作表 '{sheet_name}'，当前共 {len(wb.sheetnames)} 个"


@tool
@with_file_lock
def remove_sheet(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
) -> str:
    """删除输出文件中的指定工作表。

    Args:
        sheet_name: 要删除的工作表名称
    """
    wb = _load(state)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    del wb[sheet_name]
    _save(wb, state)
    return f"已删除工作表 '{sheet_name}'，剩余: {wb.sheetnames}"


@tool
@with_file_lock
def rename_sheet(
    state: Annotated[dict, InjectedState],
    old_name: str,
    new_name: str,
) -> str:
    """重命名输出文件中的指定工作表。

    Args:
        old_name: 当前工作表名称
        new_name: 新工作表名称
    """
    wb = _load(state)
    if old_name not in wb.sheetnames:
        wb.close()
        raise ToolException(f"工作表 '{old_name}' 不存在")
    if new_name in wb.sheetnames:
        wb.close()
        raise ToolException(f"工作表 '{new_name}' 已存在")
    wb[old_name].title = new_name
    _save(wb, state)
    return f"已将 '{old_name}' 重命名为 '{new_name}'"


@tool
@with_file_lock
def copy_sheet(
    state: Annotated[dict, InjectedState],
    source_name: str,
    target_name: str,
) -> str:
    """复制一个工作表并以新名称保存。

    Args:
        source_name: 源工作表名称
        target_name: 新工作表名称
    """
    wb = _load(state)
    if source_name not in wb.sheetnames:
        wb.close()
        raise ToolException(f"工作表 '{source_name}' 不存在")
    if target_name in wb.sheetnames:
        wb.close()
        raise ToolException(f"工作表 '{target_name}' 已存在")
    new_ws = wb.copy_worksheet(wb[source_name])
    new_ws.title = target_name
    _save(wb, state)
    return f"已复制 '{source_name}' → '{target_name}'"
