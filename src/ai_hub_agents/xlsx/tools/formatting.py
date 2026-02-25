"""格式化工具。"""

from __future__ import annotations

from typing import Annotated, Optional

from langchain_core.tools import ToolException, tool
from langgraph.prebuilt import InjectedState
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from ._helpers import ensure_dict, get_output_file, with_file_lock


def _load(state: dict):
    return load_workbook(get_output_file(state))


def _iter_cells(ws, range_str: str):
    """遍历范围内的所有单元格。"""
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col,
    ):
        yield from row


@tool
@with_file_lock
def format_cells(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    range_str: str,
    font: Optional[dict | str] = None,
    fill: Optional[dict | str] = None,
    alignment: Optional[dict | str] = None,
    number_format: Optional[str] = None,
    border: Optional[dict | str] = None,
) -> str:
    """为输出文件中指定范围的单元格应用格式。所有格式参数均可选，按需传入。

    Args:
        sheet_name: 工作表名称
        range_str: 单元格范围，如 "A1:D10"
        font: 字体，如 {"name": "Arial", "size": 12, "bold": true, "italic": false, "color": "0000FF"}
        fill: 填充，如 {"color": "FFFF00", "fill_type": "solid"}
        alignment: 对齐，如 {"horizontal": "center", "vertical": "center", "wrap_text": true}
        number_format: 数字格式，如 "$#,##0", "0.0%", "#,##0;(#,##0)"
        border: 边框，如 {"style": "thin", "color": "000000"}。style 可为 thin/medium/thick/double/dashed/dotted
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    font = ensure_dict(font)
    fill = ensure_dict(fill)
    alignment = ensure_dict(alignment)
    border = ensure_dict(border)

    output = get_output_file(state)
    try:
        wb = _load(state)
        ws = wb[sheet_name]
        count = 0

        font_obj = None
        if font:
            font_obj = Font(**font)

        fill_obj = None
        if fill:
            fill_color = fill.get("color", "FFFFFF")
            fill_type = fill.get("fill_type", "solid")
            fill_obj = PatternFill(start_color=fill_color, end_color=fill_color, fill_type=fill_type)

        align_obj = None
        if alignment:
            align_obj = Alignment(**alignment)

        border_obj = None
        if border:
            side = Side(style=border.get("style", "thin"), color=border.get("color", "000000"))
            border_obj = Border(top=side, bottom=side, left=side, right=side)

        for cell in _iter_cells(ws, range_str):
            if font_obj:
                cell.font = font_obj
            if fill_obj:
                cell.fill = fill_obj
            if align_obj:
                cell.alignment = align_obj
            if number_format:
                cell.number_format = number_format
            if border_obj:
                cell.border = border_obj
            count += 1

        wb.save(output)
        applied = []
        if font:
            applied.append("字体")
        if fill:
            applied.append("填充")
        if alignment:
            applied.append("对齐")
        if number_format:
            applied.append("数字格式")
        if border:
            applied.append("边框")
        return f"已为 {count} 个单元格 ({range_str}) 应用: {', '.join(applied)}"
    except KeyError:
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    except Exception as e:
        raise ToolException(f"format_cells 失败: {e}") from e


@tool
@with_file_lock
def set_column_width(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    columns: dict[str, float],
) -> str:
    """设置输出文件中指定列的宽度。

    Args:
        sheet_name: 工作表名称
        columns: {列字母: 宽度} 字典，如 {"A": 15.0, "B": 20.0}
    """
    output = get_output_file(state)
    try:
        wb = _load(state)
        ws = wb[sheet_name]
        for col, width in columns.items():
            ws.column_dimensions[col.upper()].width = width
        wb.save(output)
        return f"已设置 {len(columns)} 列宽度: {columns}"
    except KeyError:
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    except Exception as e:
        raise ToolException(f"set_column_width 失败: {e}") from e


@tool
@with_file_lock
def set_row_height(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    rows: dict[int, float],
) -> str:
    """设置输出文件中指定行的高度。

    Args:
        sheet_name: 工作表名称
        rows: {行号: 高度} 字典，如 {1: 25.0, 2: 18.0}
    """
    output = get_output_file(state)
    try:
        wb = _load(state)
        ws = wb[sheet_name]
        for row_num, height in rows.items():
            ws.row_dimensions[row_num].height = height
        wb.save(output)
        return f"已设置 {len(rows)} 行高度: {rows}"
    except KeyError:
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    except Exception as e:
        raise ToolException(f"set_row_height 失败: {e}") from e


@tool
@with_file_lock
def merge_cells(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    range_str: str,
) -> str:
    """合并输出文件中指定范围的单元格。

    Args:
        sheet_name: 工作表名称
        range_str: 合并范围，如 "A1:D1"
    """
    output = get_output_file(state)
    try:
        wb = _load(state)
        wb[sheet_name].merge_cells(range_str)
        wb.save(output)
        return f"已合并 {range_str}"
    except KeyError:
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    except Exception as e:
        raise ToolException(f"merge_cells 失败: {e}") from e


@tool
@with_file_lock
def unmerge_cells(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    range_str: str,
) -> str:
    """取消合并输出文件中指定范围的单元格。

    Args:
        sheet_name: 工作表名称
        range_str: 取消合并范围，如 "A1:D1"
    """
    output = get_output_file(state)
    try:
        wb = _load(state)
        wb[sheet_name].unmerge_cells(range_str)
        wb.save(output)
        return f"已取消合并 {range_str}"
    except KeyError:
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    except Exception as e:
        raise ToolException(f"unmerge_cells 失败: {e}") from e


@tool
@with_file_lock
def freeze_panes(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    cell: str,
) -> str:
    """冻结窗格。指定单元格左上方的行和列将被冻结。

    Args:
        sheet_name: 工作表名称
        cell: 冻结位置，如 "B2" 冻结第一行和第一列，"A2" 仅冻结第一行
    """
    output = get_output_file(state)
    try:
        wb = _load(state)
        wb[sheet_name].freeze_panes = cell
        wb.save(output)
        return f"已冻结窗格于 {cell}"
    except KeyError:
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    except Exception as e:
        raise ToolException(f"freeze_panes 失败: {e}") from e


@tool
@with_file_lock
def auto_fit_columns(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    columns: Optional[list[str]] = None,
    min_width: float = 8.0,
    max_width: float = 50.0,
) -> str:
    """根据内容自动调整列宽。

    Args:
        sheet_name: 工作表名称
        columns: 指定列字母列表如 ["A","B"]，不指定则调整所有列
        min_width: 最小宽度，默认 8.0
        max_width: 最大宽度，默认 50.0
    """
    from openpyxl.utils import get_column_letter

    output = get_output_file(state)
    try:
        wb = _load(state)
        ws = wb[sheet_name]
        target_cols = set()
        if columns:
            target_cols = {c.upper() for c in columns}

        adjusted = []
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if target_cols and col_letter not in target_cols:
                continue
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            width = min(max(max_len + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = width
            adjusted.append(f"{col_letter}={width:.1f}")

        wb.save(output)
        return f"已自动调整 {len(adjusted)} 列宽: {', '.join(adjusted)}"
    except KeyError:
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    except Exception as e:
        raise ToolException(f"auto_fit_columns 失败: {e}") from e
