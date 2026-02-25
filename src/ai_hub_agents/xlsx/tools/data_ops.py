"""数据操作工具。"""

from __future__ import annotations

from typing import Annotated, Optional

from langchain_core.tools import ToolException, tool
from langgraph.prebuilt import InjectedState
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from ._helpers import ensure_dict, get_output_file, with_file_lock


def _load(state: dict):
    return load_workbook(get_output_file(state))


@tool
@with_file_lock
def find_replace(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    find: str,
    replace: str,
    match_case: bool = False,
    in_range: Optional[str] = None,
) -> str:
    """在输出文件中查找并替换单元格值。

    Args:
        sheet_name: 工作表名称
        find: 查找的文本
        replace: 替换为的文本
        match_case: 是否区分大小写，默认 False
        in_range: 限定搜索范围如 "A1:D10"，不指定则搜索整个工作表
    """
    output = get_output_file(state)
    try:
        wb = _load(state)
        ws = wb[sheet_name]
        count = 0

        if in_range:
            min_col, min_row, max_col, max_row = range_boundaries(in_range)
            cells = ws.iter_rows(
                min_row=min_row, max_row=max_row,
                min_col=min_col, max_col=max_col,
            )
        else:
            cells = ws.iter_rows()

        for row in cells:
            for cell in row:
                if cell.value is None or not isinstance(cell.value, str):
                    continue
                val = cell.value
                if match_case:
                    if find in val:
                        cell.value = val.replace(find, replace)
                        count += 1
                else:
                    idx = val.lower().find(find.lower())
                    if idx != -1:
                        cell.value = val[:idx] + replace + val[idx + len(find):]
                        count += 1

        wb.save(output)
        return f"已替换 {count} 处: '{find}' → '{replace}'"
    except KeyError:
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    except Exception as e:
        raise ToolException(f"find_replace 失败: {e}") from e


@tool
@with_file_lock
def copy_range(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    source_range: str,
    target_cell: str,
    target_sheet: Optional[str] = None,
    values_only: bool = False,
) -> str:
    """复制单元格范围到另一个位置（可跨 sheet）。

    Args:
        sheet_name: 源工作表名称
        source_range: 源范围，如 "A1:D10"
        target_cell: 目标起始单元格，如 "F1"
        target_sheet: 目标工作表，不指定则在同一 sheet
        values_only: 仅复制值（不含公式），默认 False
    """
    from openpyxl.utils import column_index_from_string

    output = get_output_file(state)
    try:
        wb = _load(state)
        ws_src = wb[sheet_name]
        ws_dst = wb[target_sheet] if target_sheet else ws_src

        min_col, min_row, max_col, max_row = range_boundaries(source_range)
        dst_letter = "".join(c for c in target_cell if c.isalpha())
        dst_row = int("".join(c for c in target_cell if c.isdigit()))
        dst_col = column_index_from_string(dst_letter)

        count = 0
        for ri, row in enumerate(ws_src.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col,
        )):
            for ci, cell in enumerate(row):
                dst_cell = ws_dst.cell(row=dst_row + ri, column=dst_col + ci)
                if values_only and isinstance(cell.value, str) and cell.value.startswith("="):
                    dst_cell.value = cell.value
                else:
                    dst_cell.value = cell.value
                dst_cell.font = cell.font.copy()
                dst_cell.fill = cell.fill.copy()
                dst_cell.alignment = cell.alignment.copy()
                dst_cell.number_format = cell.number_format
                dst_cell.border = cell.border.copy()
                count += 1

        wb.save(output)
        dst_name = target_sheet or sheet_name
        return f"已复制 {count} 个单元格: {sheet_name}!{source_range} → {dst_name}!{target_cell}"
    except KeyError as e:
        raise ToolException(f"工作表不存在: {e}")
    except Exception as e:
        raise ToolException(f"copy_range 失败: {e}") from e


@tool
@with_file_lock
def sort_range(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    range_str: str,
    sort_col: int,
    ascending: bool = True,
    has_header: bool = True,
) -> str:
    """对指定范围的数据按列排序。

    Args:
        sheet_name: 工作表名称
        range_str: 数据范围，如 "A1:D10"
        sort_col: 排序列在范围内的相对位置（1-based），如第 2 列
        ascending: 升序排列，默认 True
        has_header: 是否包含表头行（排序时跳过），默认 True
    """
    output = get_output_file(state)
    try:
        wb = _load(state)
        ws = wb[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(range_str)

        data_start = min_row + 1 if has_header else min_row
        rows_data = []
        for r in range(data_start, max_row + 1):
            row_vals = []
            for c in range(min_col, max_col + 1):
                row_vals.append(ws.cell(row=r, column=c).value)
            rows_data.append(row_vals)

        sort_idx = sort_col - 1
        rows_data.sort(
            key=lambda x: (x[sort_idx] is None, x[sort_idx] if x[sort_idx] is not None else ""),
            reverse=not ascending,
        )

        for ri, row_vals in enumerate(rows_data):
            for ci, val in enumerate(row_vals):
                ws.cell(row=data_start + ri, column=min_col + ci, value=val)

        wb.save(output)
        direction = "升序" if ascending else "降序"
        return f"已按第 {sort_col} 列{direction}排序 {len(rows_data)} 行数据"
    except KeyError:
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    except Exception as e:
        raise ToolException(f"sort_range 失败: {e}") from e


@tool
@with_file_lock
def add_conditional_format(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    range_str: str,
    rule_type: str,
    operator: Optional[str] = None,
    formula: Optional[str] = None,
    threshold: Optional[float] = None,
    format: Optional[dict | str] = None,
) -> str:
    """添加条件格式规则。

    Args:
        sheet_name: 工作表名称
        range_str: 应用范围，如 "A1:A100"
        rule_type: 规则类型，支持 "cell_value", "color_scale", "data_bar", "formula"
        operator: 运算符（cell_value 时必填），如 "greaterThan", "lessThan", "equal", "between"
        formula: 公式或阈值（cell_value 时为比较值如 "100"，formula 类型时为公式如 "=$A1>0"）
        threshold: 数值阈值（cell_value 时可替代 formula）
        format: 满足条件时的格式，如 {"font_color": "FF0000", "bg_color": "FFFF00", "bold": true}
    """
    from openpyxl.formatting.rule import (
        CellIsRule,
        ColorScaleRule,
        DataBarRule,
        FormulaRule,
    )
    from openpyxl.styles import Font, PatternFill

    format = ensure_dict(format)
    output = get_output_file(state)
    try:
        wb = _load(state)
        ws = wb[sheet_name]

        if rule_type == "cell_value":
            fmt = format or {}
            font = Font(color=fmt.get("font_color"), bold=fmt.get("bold"))
            fill = PatternFill(
                start_color=fmt.get("bg_color", "FFFFFF"),
                end_color=fmt.get("bg_color", "FFFFFF"),
                fill_type="solid",
            ) if fmt.get("bg_color") else None
            val = formula or str(threshold)
            rule = CellIsRule(
                operator=operator or "greaterThan",
                formula=[val],
                font=font,
                fill=fill,
            )
        elif rule_type == "color_scale":
            rule = ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            )
        elif rule_type == "data_bar":
            rule = DataBarRule(
                start_type="min", end_type="max",
                color="638EC6",
            )
        elif rule_type == "formula":
            fmt = format or {}
            font = Font(color=fmt.get("font_color"), bold=fmt.get("bold"))
            fill = PatternFill(
                start_color=fmt.get("bg_color", "FFFFFF"),
                end_color=fmt.get("bg_color", "FFFFFF"),
                fill_type="solid",
            ) if fmt.get("bg_color") else None
            rule = FormulaRule(formula=[formula], font=font, fill=fill)
        else:
            raise ToolException(
                f"不支持的 rule_type: {rule_type}，"
                "可选: cell_value, color_scale, data_bar, formula"
            )

        ws.conditional_formatting.add(range_str, rule)
        wb.save(output)
        return f"已为 {range_str} 添加 {rule_type} 条件格式"
    except KeyError:
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    except ToolException:
        raise
    except Exception as e:
        raise ToolException(f"add_conditional_format 失败: {e}") from e


@tool
@with_file_lock
def add_data_validation(
    state: Annotated[dict, InjectedState],
    sheet_name: str,
    range_str: str,
    validation_type: str,
    formula1: Optional[str] = None,
    formula2: Optional[str] = None,
    allow_blank: bool = True,
    error_message: Optional[str] = None,
) -> str:
    """添加数据验证规则。

    Args:
        sheet_name: 工作表名称
        range_str: 应用范围，如 "B2:B100"
        validation_type: 验证类型，支持 "list", "whole", "decimal", "date", "textLength", "custom"
        formula1: 验证公式/值。list 类型为逗号分隔选项如 "Yes,No,Maybe"；
                  whole/decimal 为最小值；date 为最早日期；textLength 为最小长度
        formula2: 第二个公式/值（范围验证时的最大值）
        allow_blank: 是否允许空值，默认 True
        error_message: 验证失败时的错误提示
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    output = get_output_file(state)
    try:
        wb = _load(state)
        ws = wb[sheet_name]

        if validation_type == "list":
            dv = DataValidation(type="list", formula1=f'"{formula1}"', allow_blank=allow_blank)
        else:
            dv = DataValidation(
                type=validation_type,
                operator="between" if formula2 else "greaterThanOrEqual",
                formula1=formula1,
                formula2=formula2,
                allow_blank=allow_blank,
            )

        if error_message:
            dv.error = error_message
            dv.showErrorMessage = True

        dv.add(range_str)
        ws.add_data_validation(dv)
        wb.save(output)
        return f"已为 {range_str} 添加 {validation_type} 数据验证"
    except KeyError:
        raise ToolException(f"工作表 '{sheet_name}' 不存在")
    except Exception as e:
        raise ToolException(f"add_data_validation 失败: {e}") from e
