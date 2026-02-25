"""xlsx Agent 集成测试公共 fixture。"""

from __future__ import annotations

import pytest
from openpyxl import Workbook, load_workbook

from ai_hub_agents.test import ColorStreamRenderer, load_test_llm, setup_logging
from ai_hub_agents.xlsx import XlsxAgent


def pytest_configure(config):
    setup_logging()


@pytest.fixture(scope="session")
def agent():
    """整个测试会话复用同一个 agent 实例。"""
    llm = load_test_llm()
    return XlsxAgent.create(llm)


@pytest.fixture()
def xlsx_env(tmp_path):
    """创建带有示例数据的 xlsx 文件，返回 (input_file, output_file) 路径。"""
    input_file = tmp_path / "input.xlsx"
    output_file = tmp_path / "output.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["商品名称", "品类", "下单次数", "退货人数", "销售额"])
    ws.append(["商品1", "口红", 182, 90, 13988])
    ws.append(["商品2", "数码", 807, 0, 33935])
    ws.append(["商品3", "图书", 810, 766, 472])
    ws.append(["商品4", "食品", 186, 144, 2100])
    ws.append(["商品5", "数码", 500, 0, 25000])
    ws.append(["商品6", "口红", 320, 187, 8900])
    ws.append(["商品7", "食品", 90, 0, 1200])
    wb.save(input_file)

    return str(input_file), str(output_file)


@pytest.fixture()
def renderer():
    return ColorStreamRenderer()


def invoke(agent, task: str, xlsx_env, renderer) -> tuple[str, load_workbook]:
    """公共调用函数：执行 agent 并返回 (结果文本, 输出workbook)。"""
    input_file, output_file = xlsx_env
    result = agent.invoke(
        task,
        callbacks=[renderer],
        input_file=input_file,
        output_file=output_file,
    )
    wb = load_workbook(output_file)
    return result, wb
